#!/usr/bin/env python3
"""Génère un fichier Excel récapitulatif des programmes candidats."""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

wb = Workbook()
ws = wb.active
ws.title = "Programmes candidats"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="002395", end_color="002395", fill_type="solid")
green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
yellow_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
orange_fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
blue_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
gray_fill = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
link_font = Font(color="0563C1", underline="single")
center = Alignment(horizontal="center")

# === Onglet 1 : Détail par candidat ===

headers = ["Ville", "Candidat", "Parti", "Site de campagne", "Programme URL", "Statut", "Nb propositions", "Programme complet"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 30
ws.column_dimensions["D"].width = 40
ws.column_dimensions["E"].width = 40
ws.column_dimensions["F"].width = 20
ws.column_dimensions["G"].width = 18
ws.column_dimensions["H"].width = 18

# Charger données
with open(os.path.join(BASE, "data", "villes.json"), "r", encoding="utf-8") as f:
    villes = json.load(f)

villes.sort(key=lambda v: v["nom"])

row = 2
for v in villes:
    election_id = v["elections"][0]
    election_file = os.path.join(BASE, "data", "elections", f"{election_id}.json")
    if not os.path.exists(election_file):
        continue
    with open(election_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    candidats = data.get("candidats", [])
    if not candidats:
        continue

    for c in candidats:
        nb_props = sum(
            1 for cat in data.get("categories", [])
            for st in cat.get("sousThemes", [])
            if st.get("propositions", {}).get(c["id"]) is not None
        )

        site = c.get("siteWeb", "") or ""
        prog_url = c.get("programmeUrl", "") or ""
        prog_complet = c.get("programmeComplet", False)
        prog_pdf = c.get("programmePdfPath", "") or ""

        # Statut
        if prog_pdf and prog_pdf != "#":
            statut = "PDF telecharge"
            fill = green_fill
        elif prog_complet:
            statut = "Programme complet"
            fill = green_fill
        elif nb_props > 5:
            statut = "Bien couvert"
            fill = yellow_fill
        elif nb_props > 0:
            statut = "Partiellement couvert"
            fill = orange_fill
        elif (site and site != "#") or (prog_url and prog_url != "#"):
            statut = "Site identifie"
            fill = blue_fill
        else:
            statut = "En attente"
            fill = gray_fill

        nom_complet = f'{c.get("prenom", "")} {c.get("nom", "")}'.strip()
        site_display = site if site and site != "#" else ""
        prog_display = prog_url if prog_url and prog_url != "#" else ""

        ws.cell(row=row, column=1, value=v["nom"]).border = thin_border
        ws.cell(row=row, column=2, value=nom_complet).border = thin_border
        ws.cell(row=row, column=3, value=c.get("parti", "")).border = thin_border

        cell_site = ws.cell(row=row, column=4, value=site_display)
        cell_site.border = thin_border
        if site_display:
            cell_site.font = link_font

        cell_prog = ws.cell(row=row, column=5, value=prog_display)
        cell_prog.border = thin_border
        if prog_display:
            cell_prog.font = link_font

        cell_statut = ws.cell(row=row, column=6, value=statut)
        cell_statut.fill = fill
        cell_statut.border = thin_border
        cell_statut.alignment = center

        ws.cell(row=row, column=7, value=nb_props).border = thin_border
        ws.cell(row=row, column=7).alignment = center

        ws.cell(row=row, column=8, value="Oui" if prog_complet else "Non").border = thin_border
        ws.cell(row=row, column=8).alignment = center

        row += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{row - 1}"

# === Onglet 2 : Résumé par ville ===

ws2 = wb.create_sheet("Resume par ville")

headers2 = ["Ville", "Nb candidats", "Nb propositions total", "Nb avec site", "Nb prog complet", "Statut global"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 15
ws2.column_dimensions["C"].width = 22
ws2.column_dimensions["D"].width = 15
ws2.column_dimensions["E"].width = 18
ws2.column_dimensions["F"].width = 22

row2 = 2
for v in villes:
    election_id = v["elections"][0]
    election_file = os.path.join(BASE, "data", "elections", f"{election_id}.json")
    if not os.path.exists(election_file):
        continue
    with open(election_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    candidats = data.get("candidats", [])
    nb_cand = len(candidats)
    nb_props_total = 0
    nb_sites = 0
    nb_complets = 0

    for c in candidats:
        nb_p = sum(
            1 for cat in data.get("categories", [])
            for st in cat.get("sousThemes", [])
            if st.get("propositions", {}).get(c["id"]) is not None
        )
        nb_props_total += nb_p
        site = c.get("siteWeb", "") or ""
        if site and site != "#":
            nb_sites += 1
        if c.get("programmeComplet", False):
            nb_complets += 1

    if nb_complets > 0:
        statut = "Programme(s) complet(s)"
        fill = green_fill
    elif nb_props_total > 20:
        statut = "Bien couvert"
        fill = yellow_fill
    elif nb_props_total > 0:
        statut = "Partiellement couvert"
        fill = orange_fill
    elif nb_sites > 0:
        statut = "Sites identifies"
        fill = blue_fill
    else:
        statut = "En attente"
        fill = gray_fill

    ws2.cell(row=row2, column=1, value=v["nom"]).border = thin_border
    ws2.cell(row=row2, column=2, value=nb_cand).border = thin_border
    ws2.cell(row=row2, column=2).alignment = center
    ws2.cell(row=row2, column=3, value=nb_props_total).border = thin_border
    ws2.cell(row=row2, column=3).alignment = center
    ws2.cell(row=row2, column=4, value=nb_sites).border = thin_border
    ws2.cell(row=row2, column=4).alignment = center
    ws2.cell(row=row2, column=5, value=nb_complets).border = thin_border
    ws2.cell(row=row2, column=5).alignment = center
    cell_s = ws2.cell(row=row2, column=6, value=statut)
    cell_s.fill = fill
    cell_s.border = thin_border
    cell_s.alignment = center

    row2 += 1

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:F{row2 - 1}"

# === Onglet 3 : Guide de lecture ===

ws3 = wb.create_sheet("Guide de lecture")
ws3.sheet_properties.tabColor = "002395"

# Mettre cet onglet en premier
wb.move_sheet("Guide de lecture", offset=-2)

ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 25
ws3.column_dimensions["C"].width = 80

title_font = Font(bold=True, size=14, color="002395")
section_font = Font(bold=True, size=12, color="002395")
bold_font = Font(bold=True, size=11)
normal_font = Font(size=11)
wrap = Alignment(wrap_text=True, vertical="top")

r = 1
ws3.cell(row=r, column=2, value="Guide de lecture — Programmes candidats municipales 2026").font = title_font
ws3.merge_cells("B1:C1")
ws3.row_dimensions[r].height = 30

r = 3
ws3.cell(row=r, column=2, value="Ce fichier recense les 103 villes du top 100 + communes supplementaires").font = normal_font
ws3.merge_cells(f"B{r}:C{r}")
r = 4
ws3.cell(row=r, column=2, value="integrees au comparateur pourquituvotes.fr pour les municipales 2026.").font = normal_font
ws3.merge_cells(f"B{r}:C{r}")
r = 5
ws3.cell(row=r, column=2, value="Derniere mise a jour : 12 fevrier 2026 — Date limite candidatures : 26 fevrier 2026").font = Font(size=11, italic=True)
ws3.merge_cells(f"B{r}:C{r}")

r = 7
ws3.cell(row=r, column=2, value="ONGLETS").font = section_font
r = 9
ws3.cell(row=r, column=2, value="Programmes candidats").font = bold_font
ws3.cell(row=r, column=3, value="Detail de chaque candidat : ville, nom, parti, site web, URL programme, statut, nombre de propositions extraites, programme complet ou non.").font = normal_font
ws3.cell(row=r, column=3).alignment = wrap
ws3.row_dimensions[r].height = 35
r = 11
ws3.cell(row=r, column=2, value="Resume par ville").font = bold_font
ws3.cell(row=r, column=3, value="Vue synthetique par ville : nombre de candidats, total de propositions extraites, nombre de candidats avec un site identifie, nombre de programmes complets, et statut global.").font = normal_font
ws3.cell(row=r, column=3).alignment = wrap
ws3.row_dimensions[r].height = 35

r = 14
ws3.cell(row=r, column=2, value="COLONNES (onglet detail)").font = section_font
r = 16
colonnes = [
    ("Ville", "Nom de la commune"),
    ("Candidat", "Prenom et nom du candidat ou tete de liste"),
    ("Parti", "Etiquette politique ou coalition"),
    ("Site de campagne", "URL du site officiel de campagne (si identifie)"),
    ("Programme URL", "Lien direct vers la page programme ou le PDF en ligne"),
    ("Statut", "Niveau de couverture dans le comparateur (voir legende ci-dessous)"),
    ("Nb propositions", "Nombre de propositions deja extraites et integrees au comparateur"),
    ("Programme complet", "Oui si le programme complet a ete traite (PDF ou web), Non sinon"),
]
for colonne, desc in colonnes:
    ws3.cell(row=r, column=2, value=colonne).font = bold_font
    ws3.cell(row=r, column=3, value=desc).font = normal_font
    r += 1

r += 1
ws3.cell(row=r, column=2, value="LEGENDE DES STATUTS").font = section_font
r += 2

statuts = [
    ("PDF telecharge", green_fill, "Le PDF du programme a ete telecharge et traite. Les propositions sont integrees."),
    ("Programme complet", green_fill, "Le programme complet a ete traite (web ou PDF). Couverture maximale."),
    ("Bien couvert", yellow_fill, "Plus de 5 propositions extraites. Bonne couverture mais potentiellement incomplete."),
    ("Partiellement couvert", orange_fill, "Entre 1 et 5 propositions extraites. Donnees issues de la presse ou de bribes de programme."),
    ("Site identifie", blue_fill, "Un site de campagne ou une page programme existe mais n'a pas encore ete exploree."),
    ("En attente", gray_fill, "Aucun programme publie ni site de campagne identifie a ce jour."),
]
for statut_name, fill, desc in statuts:
    cell_s = ws3.cell(row=r, column=2, value=statut_name)
    cell_s.font = bold_font
    cell_s.fill = fill
    cell_s.border = thin_border
    cell_s.alignment = center
    cell_d = ws3.cell(row=r, column=3, value=desc)
    cell_d.font = normal_font
    cell_d.alignment = wrap
    ws3.row_dimensions[r].height = 30
    r += 1

r += 1
ws3.cell(row=r, column=2, value="PRIORITES D'ACTION").font = section_font
r += 2

actions = [
    ("1. Sites 'a explorer'", "Les candidats avec statut 'Site identifie' ont un site web mais leurs propositions n'ont pas encore ete extraites. Ce sont les plus faciles a completer."),
    ("2. Programmes en ligne", "Delogu (Marseille), Doucet (Lyon), Hurmic (Bordeaux), Dati (Paris) ont des programmes en ligne deja partiellement integres. A reverifier pour mises a jour."),
    ("3. PDF a telecharger", "Gregoire (Paris) a un PDF disponible non encore telecharge."),
    ("4. Attente publications", "La majorite des candidats n'ont pas encore publie de programme. Les programmes officiels seront disponibles d'ici debut mars 2026."),
    ("5. Verification pre-scrutin", "Faire un dernier passage sur toutes les villes debut mars avant le 1er tour (15 mars 2026)."),
]
for action, desc in actions:
    ws3.cell(row=r, column=2, value=action).font = bold_font
    cell_d = ws3.cell(row=r, column=3, value=desc)
    cell_d.font = normal_font
    cell_d.alignment = wrap
    ws3.row_dimensions[r].height = 40
    r += 1

r += 1
ws3.cell(row=r, column=2, value="REGENERER CE FICHIER").font = section_font
r += 2
ws3.cell(row=r, column=2, value="Commande :").font = bold_font
ws3.cell(row=r, column=3, value="python scripts/generer_excel_programmes.py").font = Font(size=11, name="Consolas")
r += 1
ws3.cell(row=r, column=2, value="Fichier genere :").font = bold_font
ws3.cell(row=r, column=3, value="docs/programmes-candidats-2026.xlsx").font = Font(size=11, name="Consolas")

# Sauvegarder
output = os.path.join(BASE, "docs", "programmes-candidats-2026.xlsx")
wb.save(output)
print(f"Excel genere : {output}")
print(f"Onglet 1 : Guide de lecture")
print(f"Onglet 2 : {row - 2} lignes (detail par candidat)")
print(f"Onglet 3 : {row2 - 2} lignes (resume par ville)")
